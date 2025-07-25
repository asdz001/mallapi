# shop/admin_views/admin_import.py

import openpyxl
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from shop.models import RawProduct
from django.db import transaction
from django.http import HttpResponse
import os
from django.conf import settings
from openpyxl import load_workbook
from shop.models import RawProduct, RawProductOption
from shop.models import Product, ProductOption
from shop.utils.excel_helper import generate_failed_excel
import io

@staff_member_required
@csrf_exempt
def import_rawproduct_excel(request):
    context = {}
    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file:
            context["error"] = "파일이 업로드되지 않았습니다."
            return render(request, "admin/shop/rawproduct/import_excel_form.html", context)

        try:
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            success, fail, errors = 0, 0, []
            failed_rows = []
            
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = dict(zip(header, row))

                    # ✅ 필수값 확인
                    if not row_data.get("retailer") or not row_data.get("external_product_id"):
                        raise Exception(f"{i}행: retailer 또는 external_product_id가 없습니다.")

                    # ✅ 상품 등록 또는 업데이트
                    product_obj, _ = RawProduct.objects.update_or_create(
                        retailer=row_data["retailer"],
                        external_product_id=row_data["external_product_id"],
                        defaults={
                            "raw_brand_name": row_data.get("raw_brand_name"),
                            "product_name": row_data.get("product_name"),
                            "gender": row_data.get("gender"),
                            "category1": row_data.get("category1"),
                            "category2": row_data.get("category2"),
                            "season": row_data.get("season"),
                            "sku": row_data.get("sku"),
                            "color": row_data.get("color"),
                            "origin": row_data.get("origin"),
                            "material": row_data.get("material"),
                            "image_url_1": row_data.get("image_url_1"),
                            "image_url_2": row_data.get("image_url_2"),
                            "image_url_3": row_data.get("image_url_3"),
                            "image_url_4": row_data.get("image_url_4"),
                            "price_org": row_data.get("price_org") or 0,
                            "price_supply": row_data.get("price_supply") or 0,
                            "discount_rate": row_data.get("discount_rate") or 0,
                            "price_retail": row_data.get("price_retail") or 0,
                            "description": row_data.get("description"),
                            "status": row_data.get("status") or "pending",
                        }
                    )

                    # ✅ 옵션 등록 (필수값: external_option_id, option_name)
                    if not row_data.get("external_option_id") or not row_data.get("option_name"):
                        raise Exception(f"{i}행: 옵션 필수값 누락 (external_option_id, option_name)")

                    RawProductOption.objects.update_or_create(
                        external_option_id=row_data["external_option_id"],
                        defaults={
                            "product": product_obj,
                            "option_name": row_data.get("option_name"),
                            "stock": row_data.get("stock") or 0,
                            "price": row_data.get("price") or 0,
                            "option_url": row_data.get("option_url"),
                        }
                    )

                    success += 1

                except Exception as e:
                    fail += 1
                    msg = f"{i}행 오류: {e}"
                    row_dict = dict(zip(header, row))
                    row_dict["오류 메시지"] = msg
                    failed_rows.append(row_dict)
                    errors.append(msg)


            context["success"] = f"✅ 등록 성공: {success}개 / 실패: {fail}개"
            context["errors"] = errors


            if failed_rows:
                failed_file_path = os.path.join(
                    settings.BASE_DIR, "shop", "static", "admin", "shop", "rawproduct", "last_failed.xlsx"
                )
                generate_failed_excel(failed_rows, failed_file_path)
                context["failed_file_url"] = "admin/shop/rawproduct/last_failed.xlsx"


        except Exception as e:
            context["error"] = f"처리 도중 오류 발생: {e}"        


    return render(request, "admin/shop/rawproduct/import_excel_form.html", context)



def export_rawproduct_excel(request):
    file_path = os.path.join(settings.BASE_DIR, 'shop', 'static', 'admin', 'shop', 'rawproduct', 'rawproduct_export.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=rawproduct_export.xlsx'
            return response
    else:
        return HttpResponse("파일이 존재하지 않습니다.", status=404)





#가공상품 등록
@staff_member_required
@csrf_exempt
def import_product_excel(request):
    context = {}
    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file:
            context["error"] = "파일이 업로드되지 않았습니다."
            return render(request, "admin/shop/product/import_excel_form.html", context)

        try:
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            success, fail, errors = 0, 0, []
            failed_rows = []
            
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = dict(zip(header, row))

                    # ✅ 필수값 확인
                    if not row_data.get("retailer") or not row_data.get("external_product_id"):
                        raise Exception(f"{i}행: retailer 또는 external_product_id가 없습니다.")

                    # ✅ 상품 등록 또는 업데이트
                    product_obj, _ = Product.objects.update_or_create(
                        retailer=row_data["retailer"],
                        external_product_id=row_data["external_product_id"],
                        defaults={
                            "brand_name": row_data.get("brand_name"),
                            "raw_brand_name": row_data.get("raw_brand_name"),
                            "product_name": row_data.get("product_name"),                            
                            "gender": row_data.get("gender"),
                            "category1": row_data.get("category1"),
                            "category2": row_data.get("category2"),
                            "season": row_data.get("season"),
                            "sku": row_data.get("sku"),
                            "color": row_data.get("color"),
                            "origin": row_data.get("origin"),
                            "material": row_data.get("material"),
                            "image_url_1": row_data.get("image_url_1"),
                            "image_url_2": row_data.get("image_url_2"),
                            "image_url_3": row_data.get("image_url_3"),
                            "image_url_4": row_data.get("image_url_4"),
                            "price_org": row_data.get("price_org") or 0,
                            "discount_rate": row_data.get("discount_rate") or 0,
                            "price_retail": row_data.get("price_retail") or 0,
                            "markup" : row_data.get("markup") or 0,
                            "calculated_price_krw": row_data.get("calculated_price_krw") or 0,
                            "description": row_data.get("description"),
                            "status": row_data.get("status") or "active",
                        }
                    )

                    if "price_supply" in row_data:
                        try:
                            product_obj.__dict__["price_supply"] = row_data.get("price_supply") or 0
                        except Exception as e:
                            print("가격 세팅 오류:", e)

                    # ✅ 옵션 등록 (필수값: external_option_id, option_name)
                    if not row_data.get("external_option_id") or not row_data.get("option_name"):
                        raise Exception(f"{i}행: 옵션 필수값 누락 (external_option_id, option_name)")

                    ProductOption.objects.update_or_create(
                        external_option_id=row_data["external_option_id"],
                        defaults={
                            "product": product_obj,
                            "option_name": row_data.get("option_name"),
                            "stock": row_data.get("stock") or 0,
                            "price": row_data.get("price") or 0,
                            "option_url": row_data.get("option_url"),
                        }
                    )

                    success += 1

                except Exception as e:
                    fail += 1
                    msg = f"{i}행 오류: {e}"
                    row_dict = dict(zip(header, row))
                    row_dict["오류 메시지"] = msg
                    failed_rows.append(row_dict)
                    errors.append(msg)


            context["success"] = f"✅ 등록 성공: {success}개 / 실패: {fail}개"
            context["errors"] = errors


            if failed_rows:
                failed_file_path = os.path.join(
                    settings.BASE_DIR, "shop", "static", "admin", "shop", "product", "last_failed.xlsx"
                )
                generate_failed_excel(failed_rows, failed_file_path)
                context["failed_file_url"] = "admin/shop/product/last_failed.xlsx"

                
        except Exception as e:
            context["error"] = f"처리 도중 오류 발생: {e}"        


    return render(request, "admin/shop/product/import_excel_form.html", context)



def export_product_excel(request):
    file_path = os.path.join(settings.BASE_DIR, 'shop', 'static', 'admin', 'shop', 'product', 'product_export.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=product_export.xlsx'
            return response
    else:
        return HttpResponse("파일이 존재하지 않습니다.", status=404)
